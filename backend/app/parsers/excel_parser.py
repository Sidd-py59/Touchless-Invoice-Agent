import io
from decimal import Decimal, InvalidOperation
from typing import Any
import openpyxl
from openpyxl.utils import get_column_letter

from app.parsers.base_parser import BaseParser
from app.schemas.parser import (
    NormalizedDocumentDTO,
    NormalizedEmployeeEntryDTO,
    NormalizedExtractionDTO,
)


class ExcelParser(BaseParser):
    """
    Parser for Excel (.xlsx) files.
    Dynamically identifies column index layout by scanning top rows for keywords.
    Packages extracted cells into DTOs without touching database models.
    """

    def parse(self, file_content: bytes, **kwargs) -> NormalizedDocumentDTO:
        # Load Excel workbook in read-only / data_only mode to get evaluated cell values
        wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
        ws = wb.active or wb.worksheets[0]

        client_name = kwargs.get("client_name", "Unknown Client")
        billing_year = int(kwargs.get("billing_year", 2026))
        billing_month = int(kwargs.get("billing_month", 6))

        entries = []
        raw_extractions = []

        # Read sheet metadata for global extraction logging
        raw_extractions.append(
            NormalizedExtractionDTO(
                field_name="sheet_name",
                field_value=ws.title,
                confidence=1.0,
                row_number=1,
                column_name="A",
                entity_type="Metadata",
            )
        )

        # Dynamic column indexing scanner
        header_row_idx = None
        col_mapping = {
            "employee_code": None,
            "employee_name": None,
            "working_days": None,
            "ot_hours": None,
            "leave_days": None,
            "remarks": None,
        }

        # Scan the first 15 rows to find the headers row
        for r_idx in range(1, min(16, ws.max_row + 1)):
            row_values = [
                str(ws.cell(row=r_idx, column=c).value or "").strip().lower()
                for c in range(1, ws.max_column + 1)
            ]

            # Detect header row containing employee descriptors and attendance metrics
            has_emp = any(
                "employee" in val or "emp" in val or "code" in val or "name" in val
                for val in row_values
            )
            has_metrics = any(
                "work" in val
                or "day" in val
                or "ot" in val
                or "overtime" in val
                or "leave" in val
                for val in row_values
            )

            if has_emp and has_metrics:
                header_row_idx = r_idx
                # Map column header names to their 1-indexed column indices
                for c in range(1, ws.max_column + 1):
                    val = (
                        str(ws.cell(row=r_idx, column=c).value or "")
                        .strip()
                        .lower()
                    )
                    if "code" in val or "id" in val:
                        col_mapping["employee_code"] = c
                    elif "name" in val:
                        col_mapping["employee_name"] = c
                    elif "work" in val or "day" in val:
                        col_mapping["working_days"] = c
                    elif "ot" in val or "overtime" in val or "extra" in val:
                        col_mapping["ot_hours"] = c
                    elif "leave" in val or "off" in val or "vacation" in val:
                        col_mapping["leave_days"] = c
                    elif "remark" in val or "note" in val or "comment" in val:
                        col_mapping["remarks"] = c
                break

        # Fallback layout mapping if no matching headers were discovered
        if header_row_idx is None:
            header_row_idx = 1
            col_mapping = {
                "employee_code": 1,
                "employee_name": 2,
                "working_days": 3,
                "ot_hours": 4,
                "leave_days": 5,
                "remarks": 6,
            }

        # Parse data rows starting right below header
        for r in range(header_row_idx + 1, ws.max_row + 1):
            code_cell = (
                ws.cell(row=r, column=col_mapping["employee_code"])
                if col_mapping["employee_code"]
                else None
            )
            name_cell = (
                ws.cell(row=r, column=col_mapping["employee_name"])
                if col_mapping["employee_name"]
                else None
            )

            # Skip empty lines
            if (code_cell is None or code_cell.value is None) and (
                name_cell is None or name_cell.value is None
            ):
                continue

            raw_code = str(code_cell.value).strip() if code_cell and code_cell.value is not None else None
            raw_name = str(name_cell.value).strip() if name_cell and name_cell.value is not None else None

            # Parse numeric quantities with decimal fallback
            working_days = (
                self._to_decimal(
                    ws.cell(row=r, column=col_mapping["working_days"]).value
                )
                if col_mapping["working_days"]
                else Decimal("0.00")
            )
            ot_hours = (
                self._to_decimal(
                    ws.cell(row=r, column=col_mapping["ot_hours"]).value
                )
                if col_mapping["ot_hours"]
                else Decimal("0.00")
            )
            leave_days = (
                self._to_decimal(
                    ws.cell(row=r, column=col_mapping["leave_days"]).value
                )
                if col_mapping["leave_days"]
                else Decimal("0.00")
            )

            remarks_val = (
                ws.cell(row=r, column=col_mapping["remarks"]).value
                if col_mapping["remarks"]
                else None
            )
            remarks = str(remarks_val).strip() if remarks_val is not None else None

            # Capture cell-level coordinates and values
            row_extractions = []
            for field, col_idx in col_mapping.items():
                if col_idx:
                    cell = ws.cell(row=r, column=col_idx)
                    if cell.value is not None:
                        row_extractions.append(
                            NormalizedExtractionDTO(
                                field_name=field,
                                field_value=str(cell.value),
                                confidence=1.0,
                                row_number=r,
                                column_name=get_column_letter(col_idx),
                                entity_type="EmployeeRecord",
                            )
                        )

            entries.append(
                NormalizedEmployeeEntryDTO(
                    raw_employee_code=raw_code,
                    raw_employee_name=raw_name,
                    working_days=working_days,
                    ot_hours=ot_hours,
                    leave_days=leave_days,
                    remarks=remarks,
                    confidence=1.0,
                    extractions=row_extractions,
                )
            )

        return NormalizedDocumentDTO(
            client_name=client_name,
            billing_year=billing_year,
            billing_month=billing_month,
            raw_extractions=raw_extractions,
            entries=entries,
        )

    def _to_decimal(self, val: Any) -> Decimal:
        if val is None:
            return Decimal("0.00")
        if isinstance(val, (int, float)):
            return Decimal(str(val))
        try:
            return Decimal(str(val).strip())
        except (InvalidOperation, ValueError):
            return Decimal("0.00")
