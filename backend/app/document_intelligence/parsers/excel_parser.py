import csv
import io
from openpyxl.utils import get_column_letter
import openpyxl

from app.document_intelligence.classifiers.source_detector import SourceType
from app.document_intelligence.dto.document_metadata import DocumentMetadata
from app.document_intelligence.dto.normalized_document import NormalizedDocument
from app.document_intelligence.mapper.schema_mapper import SchemaMapper
from app.document_intelligence.parsers.base_parser import BaseParser


class ExcelParser(BaseParser):
    """
    Excel document parser implementation.
    Reads binary file content, scans columns dynamically, and utilizes the
    SchemaMapper to convert rows to standard Pydantic employee models.
    """

    def supports(self, source_type: SourceType) -> bool:
        return source_type == SourceType.EXCEL

    async def parse(
        self, file_content: bytes, metadata: DocumentMetadata, **_: object
    ) -> NormalizedDocument:
        if metadata.file_name.lower().endswith(".csv") or "csv" in metadata.mime_type.lower():
            return await self._parse_csv(file_content, metadata)

        wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
        ws = wb.active or wb.worksheets[0]

        # Scan the first 15 rows to detect columns dynamically
        header_row_idx = None
        col_mapping = {}

        for r_idx in range(1, min(16, ws.max_row + 1)):
            row_values = [
                str(ws.cell(row=r_idx, column=c).value or "").strip().lower()
                for c in range(1, ws.max_column + 1)
            ]

            has_emp = any(
                "employee" in val or "emp" in val or "code" in val or "name" in val
                for val in row_values
            )
            has_metrics = any(
                "work" in val
                or "day" in val
                or "ot" in val
                or "overtime" in val
                or "leave" in val
                for val in row_values
            )

            if has_emp and has_metrics:
                header_row_idx = r_idx
                # Map column header names to their 1-indexed column indices
                for c in range(1, ws.max_column + 1):
                    val = (
                        str(ws.cell(row=r_idx, column=c).value or "")
                        .strip()
                        .lower()
                    )
                    if val:
                        col_mapping[val] = c
                break

        # Fallback column index mapping if no headers detected
        if header_row_idx is None:
            header_row_idx = 1
            col_mapping = {
                "employee code": 1,
                "employee name": 2,
                "working days": 3,
                "ot hours": 4,
                "leave days": 5,
                "remarks": 6,
            }

        employees = []
        raw_rows = []

        # Parse data rows starting right below header
        for r in range(header_row_idx + 1, ws.max_row + 1):
            # Extract row values as a raw column-name-to-cell-value dictionary
            raw_row = {}
            for col_name, col_idx in col_mapping.items():
                cell = ws.cell(row=r, column=col_idx)
                raw_row[col_name] = cell.value

            # Skip empty lines
            code_val = raw_row.get("employee code") or raw_row.get("code") or raw_row.get("id")
            name_val = raw_row.get("employee name") or raw_row.get("name")
            if code_val is None and name_val is None:
                continue

            raw_rows.append(raw_row)

            # Convert row to standard EmployeeRecord DTO using the SchemaMapper
            employee_rec = SchemaMapper.map_employee_row(raw_row, confidence=1.0)
            employees.append(employee_rec)

        SchemaMapper.apply_document_metadata(metadata, raw_rows)

        # Update metadata details
        metadata.parser = self.__class__.__name__
        metadata.parser_version = "1.0.0"

        return NormalizedDocument(
            metadata=metadata, employees=employees, reimbursements=[]
        )

    async def _parse_csv(
        self, file_content: bytes, metadata: DocumentMetadata
    ) -> NormalizedDocument:
        decoded = file_content.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(decoded))
        employees = []
        raw_rows = []

        for raw_row in reader:
            if not any(str(value or "").strip() for value in raw_row.values()):
                continue
            raw_rows.append(raw_row)
            employee_rec = SchemaMapper.map_employee_row(raw_row, confidence=1.0)
            if employee_rec.employee_code or employee_rec.employee_name:
                employees.append(employee_rec)

        SchemaMapper.apply_document_metadata(metadata, raw_rows)
        metadata.parser = self.__class__.__name__
        metadata.parser_version = "1.0.0"

        return NormalizedDocument(
            metadata=metadata, employees=employees, reimbursements=[]
        )